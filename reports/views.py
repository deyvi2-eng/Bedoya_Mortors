from django.views import View
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from sales.models import Sale
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils.timezone import localtime
from decimal import Decimal

class ExportSalesExcelView(LoginRequiredMixin, View):
    
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # ==========================================
        # 1. ESTILOS PROFESIONALES (OPENPYXL)
        # ==========================================
        title_font = Font(size=16, bold=True, color="0F172A")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal-600
        
        invoice_font = Font(bold=True, color="1E293B")
        invoice_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Slate-200
        
        detail_font = Font(color="334155")
        
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")
        alignment_right = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # ==========================================
        # 2. CAPTURAR FECHAS Y FILTRAR DATOS
        # ==========================================
        ventas = Sale.objects.select_related('customer', 'seller').prefetch_related('details__product').order_by('created_at')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        rango_fechas = "Histórico Completo"
        if start_date and end_date:
            ventas = ventas.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            rango_fechas = f"Desde: {start_date} - Hasta: {end_date}"
        elif start_date:
            ventas = ventas.filter(created_at__date__gte=start_date)
            rango_fechas = f"Desde: {start_date} en adelante"
        elif end_date:
            ventas = ventas.filter(created_at__date__lte=end_date)
            rango_fechas = f"Hasta: {end_date}"

        # ==========================================
        # 3. CABECERA DEL DOCUMENTO EXCEL
        # ==========================================
        ws.merge_cells('A1:I1')
        ws['A1'] = "REPORTE DETALLADO DE VENTAS - BEDOYA MOTORS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:I2')
        ws['A2'] = f"Período Filtrado: {rango_fechas}"
        ws['A2'].font = Font(italic=True, color="475569")
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.append([]) # Fila 3 vacía para dar respiro

        # Cabeceras de Tabla
        headers = [
            'Documento / Detalle de Ítems', 
            'Fecha Emisión', 
            'Cliente / Cédula', 
            'Vendedor / Método Pago', 
            'Estado', 
            'Cant.', 
            'P. Unitario', 
            'Subtotal Ítem', 
            'Total Factura'
        ]
        ws.append(headers)

        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border

        # ==========================================
        # 4. POBLAR DATOS AGRUPADOS Y SUMATORIAS
        # ==========================================
        gran_total_valido = Decimal('0.00')
        total_facturas_validas = 0

        for venta in ventas:
            fecha_local = localtime(venta.created_at).strftime("%Y-%m-%d %H:%M")
            estado = "ANULADA" if venta.is_voided else venta.get_status_display()
            
            cliente_str = f"{venta.customer.first_name} {venta.customer.last_name}\nCI: {venta.customer.cedula}" if venta.customer else "Consumidor Final\nCI: 9999999999999"
            vendedor_str = f"{venta.seller.get_full_name() if venta.seller else 'Administrador'}\n[{venta.get_payment_method_display()}]"
            
            # Solo sumamos al Gran Total si la factura NO está anulada
            if not venta.is_voided:
                gran_total_valido += venta.total
                total_facturas_validas += 1

            # --- A) FILA MAESTRA (FACTURA) ---
            row_factura = [
                f"Factura: {venta.invoice_number}",
                fecha_local,
                cliente_str,
                vendedor_str,
                estado,
                '', '', '', 
                float(venta.total)
            ]
            ws.append(row_factura)
            
            # Estilizar Fila Maestra
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 30 # Altura doble para los saltos de línea
            for col_num, cell in enumerate(ws[current_row], 1):
                cell.font = invoice_font
                cell.fill = invoice_fill
                cell.border = thin_border
                if col_num in [3, 4]: 
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif col_num == 9:
                    cell.alignment = alignment_center
                    cell.number_format = '"$"#,##0.00'
                else:
                    cell.alignment = alignment_center

            # --- B) FILAS ESCLAVAS (DETALLES) ---
            for detalle in venta.details.all():
                descripcion = detalle.service_description if detalle.is_service else (detalle.product.name if detalle.product else "Producto Eliminado")
                tipo_item = "[Servicio]" if detalle.is_service else "[Producto]"
                
                row_detalle = [
                    f"    ↳ {tipo_item} {descripcion}", # Sangría visual para indicar que pertenece a la factura superior
                    '', '', '', '', 
                    float(detalle.quantity),
                    float(detalle.unit_price),
                    float(detalle.subtotal),
                    ''
                ]
                ws.append(row_detalle)
                
                # Estilizar Filas de Detalle
                detail_row_idx = ws.max_row
                for col_num, cell in enumerate(ws[detail_row_idx], 1):
                    cell.font = detail_font
                    cell.border = thin_border
                    if col_num == 1:
                        cell.alignment = alignment_left
                    elif col_num >= 6:
                        cell.alignment = alignment_right
                        if col_num > 6:
                            cell.number_format = '"$"#,##0.00'

        # ==========================================
        # 5. FILA DE TOTAL GENERAL (FOOTER)
        # ==========================================
        ws.append([]) # Espacio en blanco antes del total
        
        total_row = [
            f'TOTAL GENERAL DE INGRESOS ({total_facturas_validas} ventas válidas)', 
            '', '', '', '', '', '', '', 
            float(gran_total_valido)
        ]
        ws.append(total_row)
        
        max_r = ws.max_row
        ws.merge_cells(start_row=max_r, start_column=1, end_row=max_r, end_column=8) # Combinar celdas para el título del total
        
        total_font = Font(bold=True, size=14, color="FFFFFF")
        total_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate-900 oscuro
        
        for col_num in range(1, 10):
            cell = ws.cell(row=max_r, column=col_num)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            if col_num == 1:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_num == 9:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '"$"#,##0.00'

        # ==========================================
        # 6. CONFIGURAR ANCHOS DE COLUMNA ESTÁTICOS
        # ==========================================
        column_widths = {
            'A': 45, # Documento / Ítems (Necesita espacio para las sangrías)
            'B': 18, # Fecha
            'C': 35, # Cliente (Necesita espacio para el salto de línea)
            'D': 30, # Vendedor
            'E': 15, # Estado
            'F': 10, # Cant
            'G': 15, # PU
            'H': 18, # Subtotal
            'I': 20, # Total Factura
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # ==========================================
        # 7. RESPUESTA HTTP
        # ==========================================
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Profesional_Ventas_BedoyaMotors.xlsx"'
        wb.save(response)

        return response